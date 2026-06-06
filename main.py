import json
import logging
import os
import re
import sqlite3
import sys
import time
import warnings
from datetime import datetime, timezone
from pathlib import Path
from tempfile import TemporaryDirectory

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage, ImageOps
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    InputMediaDocument,
    InputMediaPhoto,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)


load_dotenv(encoding="utf-8-sig")
warnings.filterwarnings("ignore", message=r"If 'per_message=False'.*", category=UserWarning)

DEFAULT_BOT_TOKEN = "8405546234:AAHwIx4gxRxtc-erFp7tckvNLNRlNQjYFV8"
DEFAULT_ADMIN_IDS = "7903688837"
DEFAULT_ADMIN_CHAT_ID = "7903688837"
DEFAULT_PUBLISH_CHAT_ID = "-1003994819171"
DEFAULT_DB_FILE = "bot.db"
DEFAULT_EXCEL_FILE = "applications.xlsx"
DEFAULT_REQUIRED_CHANNEL = "@fariks01"
DEFAULT_REQUIRED_CHANNEL_URL = "https://t.me/fariks01"

BOT_TOKEN = os.getenv("BOT_TOKEN") or DEFAULT_BOT_TOKEN
ADMIN_IDS_TEXT = os.getenv("ADMIN_IDS") or DEFAULT_ADMIN_IDS
ADMIN_CHAT_ID = os.getenv("ADMIN_CHAT_ID") or DEFAULT_ADMIN_CHAT_ID
PUBLISH_CHAT_ID = os.getenv("PUBLISH_CHAT_ID") or DEFAULT_PUBLISH_CHAT_ID or ADMIN_CHAT_ID
WEBHOOK_URL = (os.getenv("WEBHOOK_URL") or "").strip()
WEBHOOK_PATH = (os.getenv("WEBHOOK_PATH") or "telegram").strip("/") or "telegram"
RAILWAY_PUBLIC_DOMAIN = os.getenv("RAILWAY_PUBLIC_DOMAIN")
PORT = int(os.getenv("PORT") or "8080")
ADMIN_IDS = {
    int(admin_id.strip())
    for admin_id in ADMIN_IDS_TEXT.split(",")
    if admin_id.strip().lstrip("-").isdigit()
}
DB_FILE = Path(os.getenv("DB_FILE") or DEFAULT_DB_FILE)
EXCEL_FILE = Path(os.getenv("EXCEL_FILE") or DEFAULT_EXCEL_FILE)
REQUIRED_CHANNEL = os.getenv("REQUIRED_CHANNEL") or DEFAULT_REQUIRED_CHANNEL
REQUIRED_CHANNEL_URL = os.getenv("REQUIRED_CHANNEL_URL") or DEFAULT_REQUIRED_CHANNEL_URL

(
    FULL_NAME,
    BIRTH_DATE,
    ADDRESS,
    BRANCH,
    EDUCATION,
    JOB_DIRECTION,
    SPECIALTY,
    EXPERIENCE_CHOICE,
    EXPERIENCE_YEARS,
    CERTIFICATE_CHOICE,
    CERTIFICATE_PHOTO,
    PREVIOUS_JOB,
    CONVICTED,
    FAMILY_STATUS,
    PREVIOUS_SALARY,
    EXPECTED_SALARY,
    WORD_LEVEL,
    EXCEL_LEVEL,
    LANGUAGES,
    FARIKS_DURATION,
    MOTIVATION,
    PHONE,
    RECENT_PHOTO,
    REVIEW_APPLICATION,
    EDIT_FIELD,
) = range(25)
ADD_ADMIN_TARGET, REMOVE_ADMIN_TARGET, SEARCH_APPLICATION_TARGET, FILTER_BRANCH_TARGET = range(100, 104)

EDUCATION_LABELS = {
    "higher": "Oliy ma'lumotli",
    "secondary": "O'rta maxsus",
}
LEVEL_LABELS = {
    "unknown": "Bilmayman",
    "basic": "Bazaviy",
    "medium": "O'rtacha",
    "good": "Yaxshi",
}
YES_NO_LABELS = {
    "yes": "Ha",
    "no": "Yo'q",
}
JOB_DIRECTION_LABELS = {
    "admin": "Admin",
    "teacher": "O'qituvchi",
    "assistant": "O'qituvchi yordamchi",
}
SPECIALTY_LABELS = {
    "math": "Matematika",
    "physics": "Fizika",
    "russian": "Rus tili",
    "english": "Ingliz tili",
}

EDUCATION_KEYBOARD = InlineKeyboardMarkup(
    [
        [InlineKeyboardButton("🎓 Oliy ma'lumotli", callback_data="education:higher")],
        [InlineKeyboardButton("🏥 O'rta maxsus", callback_data="education:secondary")],
    ]
)
JOB_DIRECTION_KEYBOARD = InlineKeyboardMarkup(
    [
        [InlineKeyboardButton("🛠 Admin", callback_data="job:admin")],
        [InlineKeyboardButton("👨‍🏫 O'qituvchi", callback_data="job:teacher")],
        [InlineKeyboardButton("🤝 O'qituvchi yordamchi", callback_data="job:assistant")],
    ]
)
SPECIALTY_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("📐 Matematika", callback_data="specialty:math"),
            InlineKeyboardButton("⚛️ Fizika", callback_data="specialty:physics"),
        ],
        [
            InlineKeyboardButton("🇷🇺 Rus tili", callback_data="specialty:russian"),
            InlineKeyboardButton("🇬🇧 Ingliz tili", callback_data="specialty:english"),
        ],
    ]
)
EXPERIENCE_INLINE_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("✅ Ha", callback_data="exp:yes"),
            InlineKeyboardButton("❌ Yo'q", callback_data="exp:no"),
        ]
    ]
)
CERTIFICATE_CHOICE_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("✅ Bor", callback_data="cert:yes"),
            InlineKeyboardButton("❌ Yo'q", callback_data="cert:no"),
        ]
    ]
)
CONVICTED_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("✅ Ha", callback_data="convicted:yes"),
            InlineKeyboardButton("❌ Yo'q", callback_data="convicted:no"),
        ]
    ]
)
WORD_LEVEL_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("❌ Bilmayman", callback_data="word:unknown"),
            InlineKeyboardButton("🔹 Bazaviy", callback_data="word:basic"),
        ],
        [
            InlineKeyboardButton("🔸 O'rtacha", callback_data="word:medium"),
            InlineKeyboardButton("✅ Yaxshi", callback_data="word:good"),
        ],
    ]
)
EXCEL_LEVEL_KEYBOARD = InlineKeyboardMarkup(
    [
        [
            InlineKeyboardButton("❌ Bilmayman", callback_data="excel:unknown"),
            InlineKeyboardButton("🔹 Bazaviy", callback_data="excel:basic"),
        ],
        [
            InlineKeyboardButton("🔸 O'rtacha", callback_data="excel:medium"),
            InlineKeyboardButton("✅ Yaxshi", callback_data="excel:good"),
        ],
    ]
)
ADMIN_MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [["📝 Ariza yuborish"], ["🛠 Admin panel"]],
    resize_keyboard=True,
)

admin_pending_actions: dict[int, int] = {}
admin_reject_targets: dict[int, int] = {}
last_conflict_warning_at = 0.0

phone_pattern = re.compile(r"^\+?\d[\d\s()\-]{6,}$")

logging.basicConfig(
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%H:%M:%S",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
logging.getLogger("telegram.ext").setLevel(logging.WARNING)
logging.getLogger("apscheduler").setLevel(logging.WARNING)


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def db_connect() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_FILE)
    connection.row_factory = sqlite3.Row
    return connection


def init_db() -> None:
    with db_connect() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                username_lower TEXT,
                first_name TEXT,
                last_name TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS admins (
                user_id INTEGER PRIMARY KEY,
                username TEXT,
                username_lower TEXT,
                added_by INTEGER,
                added_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS applications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                username TEXT,
                full_name TEXT NOT NULL,
                birth_date TEXT NOT NULL DEFAULT '',
                address TEXT NOT NULL DEFAULT '',
                branch TEXT NOT NULL DEFAULT '',
                education TEXT NOT NULL DEFAULT '',
                job_direction TEXT NOT NULL DEFAULT '',
                specialty TEXT NOT NULL DEFAULT '',
                phone TEXT NOT NULL,
                experience TEXT NOT NULL,
                certificate_status TEXT NOT NULL DEFAULT '',
                certificate_json TEXT NOT NULL DEFAULT '{}',
                previous_job TEXT NOT NULL DEFAULT '',
                convicted TEXT NOT NULL DEFAULT '',
                family_status TEXT NOT NULL DEFAULT '',
                previous_salary TEXT NOT NULL DEFAULT '',
                expected_salary TEXT NOT NULL DEFAULT '',
                word_level TEXT NOT NULL DEFAULT '',
                excel_level TEXT NOT NULL DEFAULT '',
                languages TEXT NOT NULL DEFAULT '',
                fariks_duration TEXT NOT NULL DEFAULT '',
                motivation TEXT NOT NULL DEFAULT '',
                recent_photo_json TEXT NOT NULL DEFAULT '{}',
                role TEXT NOT NULL DEFAULT 'Admin',
                age TEXT NOT NULL DEFAULT '',
                direction TEXT NOT NULL DEFAULT '',
                certificates_json TEXT NOT NULL DEFAULT '[]',
                certificate_count INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL,
                reject_reason TEXT NOT NULL DEFAULT '',
                admin_messages_json TEXT NOT NULL DEFAULT '[]',
                created_at TEXT NOT NULL,
                decided_at TEXT,
                decided_by INTEGER
            )
            """
        )
        ensure_columns(connection, "users", {"username_lower": "TEXT"})
        ensure_columns(connection, "admins", {"username_lower": "TEXT"})
        ensure_columns(
            connection,
            "applications",
            {
                "birth_date": "TEXT NOT NULL DEFAULT ''",
                "address": "TEXT NOT NULL DEFAULT ''",
                "branch": "TEXT NOT NULL DEFAULT ''",
                "education": "TEXT NOT NULL DEFAULT ''",
                "job_direction": "TEXT NOT NULL DEFAULT ''",
                "specialty": "TEXT NOT NULL DEFAULT ''",
                "certificate_status": "TEXT NOT NULL DEFAULT ''",
                "certificate_json": "TEXT NOT NULL DEFAULT '{}'",
                "previous_job": "TEXT NOT NULL DEFAULT ''",
                "convicted": "TEXT NOT NULL DEFAULT ''",
                "family_status": "TEXT NOT NULL DEFAULT ''",
                "previous_salary": "TEXT NOT NULL DEFAULT ''",
                "expected_salary": "TEXT NOT NULL DEFAULT ''",
                "word_level": "TEXT NOT NULL DEFAULT ''",
                "excel_level": "TEXT NOT NULL DEFAULT ''",
                "fariks_duration": "TEXT NOT NULL DEFAULT ''",
                "motivation": "TEXT NOT NULL DEFAULT ''",
                "recent_photo_json": "TEXT NOT NULL DEFAULT '{}'",
                "role": "TEXT NOT NULL DEFAULT 'Admin'",
                "age": "TEXT NOT NULL DEFAULT ''",
                "direction": "TEXT NOT NULL DEFAULT ''",
                "certificates_json": "TEXT NOT NULL DEFAULT '[]'",
                "certificate_count": "INTEGER NOT NULL DEFAULT 0",
                "reject_reason": "TEXT NOT NULL DEFAULT ''",
                "admin_messages_json": "TEXT NOT NULL DEFAULT '[]'",
            },
        )
        connection.execute("CREATE INDEX IF NOT EXISTS idx_users_username_lower ON users(username_lower)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_admins_username_lower ON admins(username_lower)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_applications_user_id ON applications(user_id)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_applications_status ON applications(status)")


def ensure_columns(
    connection: sqlite3.Connection,
    table_name: str,
    columns: dict[str, str],
) -> None:
    existing_columns = {
        row["name"] for row in connection.execute(f"PRAGMA table_info({table_name})")
    }
    for column_name, column_type in columns.items():
        if column_name not in existing_columns:
            connection.execute(
                f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"
            )


def upsert_user(user) -> None:
    if not user:
        return
    with db_connect() as connection:
        connection.execute(
            """
            INSERT INTO users (
                user_id, username, username_lower, first_name, last_name, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                username = excluded.username,
                username_lower = excluded.username_lower,
                first_name = excluded.first_name,
                last_name = excluded.last_name,
                updated_at = excluded.updated_at
            """,
            (
                user.id,
                user.username,
                user.username.lower() if user.username else None,
                user.first_name,
                user.last_name,
                now_iso(),
            ),
        )


def is_admin(user_id: int | None) -> bool:
    if not user_id:
        return False
    if user_id in ADMIN_IDS:
        return True

    with db_connect() as connection:
        row = connection.execute(
            "SELECT 1 FROM admins WHERE user_id = ?",
            (user_id,),
        ).fetchone()
    return row is not None


def get_stats() -> dict:
    with db_connect() as connection:
        users_count = connection.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        db_admin_ids = {
            row["user_id"] for row in connection.execute("SELECT user_id FROM admins")
        }
        pending = connection.execute(
            "SELECT COUNT(*) FROM applications WHERE status = 'pending'"
        ).fetchone()[0]
        approved = connection.execute(
            "SELECT COUNT(*) FROM applications WHERE status = 'approved'"
        ).fetchone()[0]
        rejected = connection.execute(
            "SELECT COUNT(*) FROM applications WHERE status = 'rejected'"
        ).fetchone()[0]

    return {
        "users": users_count,
        "admins": len(db_admin_ids | ADMIN_IDS),
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
    }


def get_db_admin_ids() -> set[int]:
    with db_connect() as connection:
        return {row["user_id"] for row in connection.execute("SELECT user_id FROM admins")}


def get_all_admin_ids() -> set[int]:
    return ADMIN_IDS | get_db_admin_ids()


def get_admin_delivery_chat_ids() -> list[str]:
    chat_ids: list[str] = []
    if ADMIN_CHAT_ID:
        chat_ids.append(str(ADMIN_CHAT_ID))

    for admin_id in sorted(get_all_admin_ids()):
        admin_chat_id = str(admin_id)
        if admin_chat_id not in chat_ids:
            chat_ids.append(admin_chat_id)

    return chat_ids


def find_user_by_username(username: str) -> dict | None:
    clean_username = username.lstrip("@").lower()
    with db_connect() as connection:
        return connection.execute(
            "SELECT * FROM users WHERE username_lower = ?",
            (clean_username,),
        ).fetchone()


def find_user_by_id(user_id: int) -> dict | None:
    with db_connect() as connection:
        return connection.execute(
            "SELECT * FROM users WHERE user_id = ?",
            (user_id,),
        ).fetchone()


def add_admin(user_id: int, username: str | None, added_by: int) -> None:
    with db_connect() as connection:
        connection.execute(
            """
            INSERT INTO admins (user_id, username, username_lower, added_by, added_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(user_id) DO UPDATE SET
                username = excluded.username,
                username_lower = excluded.username_lower,
                added_by = excluded.added_by,
                added_at = excluded.added_at
            """,
            (
                user_id,
                username,
                username.lower() if username else None,
                added_by,
                now_iso(),
            ),
        )


def remove_admin(user_id: int) -> bool:
    with db_connect() as connection:
        cursor = connection.execute("DELETE FROM admins WHERE user_id = ?", (user_id,))
        return cursor.rowcount > 0


def resolve_admin_target(text: str) -> tuple[int | None, str | None, str | None]:
    target = text.strip()
    if target.lstrip("-").isdigit():
        user_id = int(target)
        known_user = find_user_by_id(user_id)
        username = known_user.get("username") if known_user else None
        return user_id, username, None

    if target.startswith("@"):
        known_user = find_user_by_username(target)
        if known_user:
            return known_user["user_id"], known_user.get("username"), None
        return (
            None,
            None,
            "Bu username bot bazasida topilmadi. Avval u foydalanuvchi botga /start bossin yoki Telegram ID yuboring.",
        )

    return None, None, "Username @ bilan yoki Telegram ID raqam ko'rinishida yuboring."


def get_application(application_id: int) -> sqlite3.Row | None:
    with db_connect() as connection:
        return connection.execute(
            "SELECT * FROM applications WHERE id = ?",
            (application_id,),
        ).fetchone()


def get_all_application_rows() -> list[sqlite3.Row]:
    with db_connect() as connection:
        return list(connection.execute("SELECT * FROM applications ORDER BY id"))


def get_approved_application_rows() -> list[sqlite3.Row]:
    with db_connect() as connection:
        return list(
            connection.execute(
                "SELECT * FROM applications WHERE status = 'approved' ORDER BY id"
            )
        )


def get_application_rows_by_status(status: str, limit: int = 10) -> list[sqlite3.Row]:
    with db_connect() as connection:
        return list(
            connection.execute(
                "SELECT * FROM applications WHERE status = ? ORDER BY id DESC LIMIT ?",
                (status, limit),
            )
        )


def search_application_rows(query: str, limit: int = 10) -> list[sqlite3.Row]:
    like_query = f"%{query.strip()}%"
    with db_connect() as connection:
        return list(
            connection.execute(
                """
                SELECT * FROM applications
                WHERE full_name LIKE ?
                   OR phone LIKE ?
                   OR branch LIKE ?
                   OR username LIKE ?
                   OR job_direction LIKE ?
                   OR specialty LIKE ?
                ORDER BY id DESC
                LIMIT ?
                """,
                (like_query, like_query, like_query, like_query, like_query, like_query, limit),
            )
        )


def get_application_rows_by_branch(branch_query: str, limit: int = 10) -> list[sqlite3.Row]:
    like_query = f"%{branch_query.strip()}%"
    with db_connect() as connection:
        return list(
            connection.execute(
                """
                SELECT * FROM applications
                WHERE branch LIKE ?
                ORDER BY id DESC
                LIMIT ?
                """,
                (like_query, limit),
            )
        )


def create_application(user, user_data: dict) -> int:
    photo = user_data.get("recent_photo", {})
    certificate = user_data.get("certificate_image", {})
    with db_connect() as connection:
        cursor = connection.execute(
            """
            INSERT INTO applications (
                user_id, username, full_name, birth_date, address, branch, education,
                job_direction, specialty, phone, experience, certificate_status,
                certificate_json, previous_job, convicted, family_status,
                previous_salary, expected_salary, word_level, excel_level, languages,
                fariks_duration, motivation, recent_photo_json, role, age, direction,
                certificates_json, certificate_count, status, created_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '[]', 0, 'pending', ?)
            """,
            (
                user.id,
                user.username,
                user_data["full_name"],
                user_data["birth_date"],
                user_data["address"],
                user_data["branch"],
                user_data["education"],
                user_data["job_direction"],
                user_data.get("specialty", ""),
                user_data["phone"],
                user_data["experience"],
                user_data.get("certificate_status", ""),
                json.dumps(certificate, ensure_ascii=False),
                user_data["previous_job"],
                user_data["convicted"],
                user_data["family_status"],
                user_data["previous_salary"],
                user_data["expected_salary"],
                user_data["word_level"],
                user_data["excel_level"],
                user_data["languages"],
                user_data["fariks_duration"],
                user_data["motivation"],
                json.dumps(photo, ensure_ascii=False),
                user_data["job_direction"],
                user_data.get("age", ""),
                user_data["branch"],
                now_iso(),
            ),
        )
        application_id = cursor.lastrowid

    safe_sync_excel_file()
    return application_id


def delete_application(application_id: int) -> None:
    with db_connect() as connection:
        connection.execute("DELETE FROM applications WHERE id = ?", (application_id,))
    safe_sync_excel_file()


def set_application_status(
    application_id: int,
    status: str,
    admin_id: int,
    reject_reason: str = "",
) -> None:
    decided_at = None if status == "pending" else now_iso()
    decided_by = None if status == "pending" else admin_id
    with db_connect() as connection:
        connection.execute(
            """
            UPDATE applications
            SET status = ?, decided_at = ?, decided_by = ?, reject_reason = ?
            WHERE id = ?
            """,
            (status, decided_at, decided_by, reject_reason, application_id),
        )
    safe_sync_excel_file()


def try_update_application_status(
    application_id: int,
    expected_status: str,
    status: str,
    admin_id: int,
    reject_reason: str = "",
) -> bool:
    with db_connect() as connection:
        cursor = connection.execute(
            """
            UPDATE applications
            SET status = ?, decided_at = ?, decided_by = ?, reject_reason = ?
            WHERE id = ? AND status = ?
            """,
            (status, now_iso(), admin_id, reject_reason, application_id, expected_status),
        )
        updated = cursor.rowcount > 0

    if updated:
        safe_sync_excel_file()
    return updated


def save_application_admin_messages(application_id: int, messages: list[dict]) -> None:
    with db_connect() as connection:
        connection.execute(
            "UPDATE applications SET admin_messages_json = ? WHERE id = ?",
            (json.dumps(messages, ensure_ascii=False), application_id),
        )


def get_application_admin_messages(application_id: int) -> list[dict]:
    row = get_application(application_id)
    if not row:
        return []
    row_data = dict(row)
    try:
        return json.loads(row_data.get("admin_messages_json") or "[]")
    except json.JSONDecodeError:
        return []


def normalize_job_direction(row_data: dict) -> str:
    raw_direction = (row_data.get("job_direction") or "").strip()
    raw_role = (row_data.get("role") or "").strip()
    valid_directions = set(JOB_DIRECTION_LABELS.values())

    if raw_direction in valid_directions:
        return raw_direction

    role_map = {
        "admin": "Admin",
        "Admin": "Admin",
        "employee": "Admin",
        "teacher": "O'qituvchi",
        "O'qituvchi": "O'qituvchi",
        "assistant": "O'qituvchi yordamchi",
        "O'qituvchi yordamchi": "O'qituvchi yordamchi",
    }
    if raw_role in role_map:
        return role_map[raw_role]

    if raw_direction:
        return role_map.get(raw_direction, raw_direction)
    return "Admin"


def row_to_application(row: sqlite3.Row | dict) -> dict:
    row_data = dict(row)
    return {
        "id": row_data["id"],
        "user_id": row_data["user_id"],
        "username": row_data.get("username"),
        "full_name": row_data.get("full_name", ""),
        "birth_date": row_data.get("birth_date", ""),
        "age": row_data.get("age", ""),
        "address": row_data.get("address", ""),
        "branch": row_data.get("branch", "") or row_data.get("direction", ""),
        "education": row_data.get("education", ""),
        "job_direction": normalize_job_direction(row_data),
        "specialty": row_data.get("specialty", ""),
        "phone": row_data.get("phone", ""),
        "experience": row_data.get("experience", ""),
        "certificate_status": row_data.get("certificate_status", ""),
        "certificate_image": json.loads(row_data.get("certificate_json") or "{}"),
        "previous_job": row_data.get("previous_job", ""),
        "convicted": row_data.get("convicted", ""),
        "family_status": row_data.get("family_status", ""),
        "previous_salary": row_data.get("previous_salary", ""),
        "expected_salary": row_data.get("expected_salary", ""),
        "word_level": row_data.get("word_level", ""),
        "excel_level": row_data.get("excel_level", ""),
        "languages": row_data.get("languages", ""),
        "fariks_duration": row_data.get("fariks_duration", ""),
        "motivation": row_data.get("motivation", ""),
        "recent_photo": json.loads(row_data.get("recent_photo_json") or "{}"),
        "status": row_data.get("status", ""),
        "reject_reason": row_data.get("reject_reason", ""),
        "created_at": row_data.get("created_at", ""),
        "decided_at": row_data.get("decided_at", ""),
        "decided_by": row_data.get("decided_by", ""),
    }


def main_keyboard_for(user_id: int) -> ReplyKeyboardMarkup | ReplyKeyboardRemove:
    if is_admin(user_id):
        return ADMIN_MAIN_KEYBOARD
    return ReplyKeyboardRemove()


def build_subscription_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("📢 Kanalga obuna bo'lish", url=REQUIRED_CHANNEL_URL)],
            [InlineKeyboardButton("✅ Obunani tekshirish", callback_data="sub:check")],
        ]
    )


async def is_user_subscribed(context: ContextTypes.DEFAULT_TYPE, user_id: int) -> bool:
    if not REQUIRED_CHANNEL:
        return True

    try:
        member = await context.bot.get_chat_member(REQUIRED_CHANNEL, user_id)
    except Exception as error:
        logger.warning("Obunani tekshirishda xatolik: %s", format_runtime_error(error))
        return False

    return member.status in {"creator", "administrator", "member"}


async def prompt_subscription(message) -> None:
    await message.reply_text(
        "📢 Ariza topshirish uchun avval Fariks kanaliga obuna bo'ling.\n"
        "Obuna bo'lgach, pastdagi \"✅ Obunani tekshirish\" tugmasini bosing.",
        reply_markup=build_subscription_keyboard(),
    )


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    upsert_user(update.effective_user)

    if is_admin(update.effective_user.id):
        await update.message.reply_text(
            "🛠 Admin menyu ochildi.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )
        return ConversationHandler.END

    if not await is_user_subscribed(context, update.effective_user.id):
        await prompt_subscription(update.message)
        return ConversationHandler.END

    return await launch_application(update.message, context)


async def begin_application(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    upsert_user(update.effective_user)
    if not is_admin(update.effective_user.id):
        if not await is_user_subscribed(context, update.effective_user.id):
            await prompt_subscription(update.message)
            return ConversationHandler.END

    return await launch_application(update.message, context)


async def subscription_check_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    upsert_user(query.from_user)
    await query.answer()

    if not await is_user_subscribed(context, query.from_user.id):
        await query.message.reply_text(
            "Hali kanalga obuna bo'lmagansiz. Iltimos, avval kanalga obuna bo'ling.",
            reply_markup=build_subscription_keyboard(),
        )
        return ConversationHandler.END

    await query.edit_message_reply_markup(reply_markup=None)
    return await launch_application(query.message, context)


async def launch_application(message, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()

    await message.reply_text(
        "Assalomu alaykum! 👋\n"
        "Fariks jamoasiga ishga ariza topshirish uchun savollarga javob bering.",
        reply_markup=ReplyKeyboardRemove(),
    )
    await message.reply_text("1. 👤 Ism-sharifingizni yozing.")
    return FULL_NAME


async def full_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    name = update.message.text.strip()
    if len(name) < 3:
        await update.message.reply_text("Iltimos, ism-sharifingizni to'liq yozing.")
        return FULL_NAME

    context.user_data["full_name"] = name
    await update.message.reply_text(
        "2. 🎂 Tug'ilgan kun, oy va yilingizni yozing.\nMasalan: 29.10.2000"
    )
    return BIRTH_DATE


async def birth_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if not is_valid_birth_date(value):
        await update.message.reply_text("Sanani to'g'ri yozing. Masalan: 29.10.2000")
        return BIRTH_DATE

    context.user_data["birth_date"] = value
    context.user_data["age"] = str(calculate_age(value))
    await update.message.reply_text("3. 📍 Yashash manzilingizni to'liq yozing.")
    return ADDRESS


def is_valid_birth_date(value: str) -> bool:
    try:
        parsed = datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        return False
    return datetime(1940, 1, 1) <= parsed <= datetime.now()


def calculate_age(birth_date_text: str) -> int:
    birth_date_value = datetime.strptime(birth_date_text, "%d.%m.%Y").date()
    today = datetime.now().date()
    age = today.year - birth_date_value.year
    if (today.month, today.day) < (birth_date_value.month, birth_date_value.day):
        age -= 1
    return age


async def address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 5:
        await update.message.reply_text("Iltimos, manzilingizni to'liqroq yozing.")
        return ADDRESS

    context.user_data["address"] = value
    await update.message.reply_text(
        "4. 🏢 Ishlashni xohlagan filialingizni yozing.\nShahar yoki tuman nomini kiriting."
    )
    return BRANCH


async def branch(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Iltimos, filial shahar yoki tuman nomini yozing.")
        return BRANCH

    context.user_data["branch"] = value
    await update.message.reply_text(
        "5. 🎓 Ma'lumot darajangizni tanlang.",
        reply_markup=EDUCATION_KEYBOARD,
    )
    return EDUCATION


async def education_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["education"] = EDUCATION_LABELS[key]
    await query.edit_message_text(f"5. 🎓 Ma'lumot: {EDUCATION_LABELS[key]}")
    await query.message.reply_text(
        "6. 🚩 Yo'nalishingizni belgilang.",
        reply_markup=JOB_DIRECTION_KEYBOARD,
    )
    return JOB_DIRECTION


async def invalid_education(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, ma'lumot darajasini tugmalardan tanlang.",
        reply_markup=EDUCATION_KEYBOARD,
    )
    return EDUCATION


async def job_direction_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["job_direction"] = JOB_DIRECTION_LABELS[key]
    context.user_data["job_direction_key"] = key
    await query.edit_message_text(f"6. 🚩 Yo'nalish: {JOB_DIRECTION_LABELS[key]}")

    if key == "admin":
        context.user_data["specialty"] = ""
        context.user_data["certificate_status"] = ""
        context.user_data["certificate_image"] = {}
        await ask_experience(query.message)
        return EXPERIENCE_CHOICE

    await query.message.reply_text(
        "7. 📚 Mutaxassisligingizni tanlang.",
        reply_markup=SPECIALTY_KEYBOARD,
    )
    return SPECIALTY


async def invalid_job_direction(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, yo'nalishni tugmalardan tanlang.",
        reply_markup=JOB_DIRECTION_KEYBOARD,
    )
    return JOB_DIRECTION


async def specialty_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["specialty"] = SPECIALTY_LABELS[key]
    await query.edit_message_text(f"7. 📚 Mutaxassislik: {SPECIALTY_LABELS[key]}")
    await query.message.reply_text("8. 💼 Ushbu yo'nalishda tajribangiz necha yil?")
    return EXPERIENCE_YEARS


async def invalid_specialty(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, mutaxassislikni tugmalardan tanlang.",
        reply_markup=SPECIALTY_KEYBOARD,
    )
    return SPECIALTY


async def ask_experience(message) -> None:
    await message.reply_text(
        "6. 💼 Shu sohada umumiy ish tajribangiz bormi?",
        reply_markup=EXPERIENCE_INLINE_KEYBOARD,
    )


async def experience_choice_callback(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    await query.edit_message_reply_markup(reply_markup=None)
    if key == "yes":
        await query.message.reply_text("⏳ Shu sohada umumiy ish tajribangiz necha yil?")
        return EXPERIENCE_YEARS

    context.user_data["experience"] = "0 yil"
    await query.message.reply_text("6. 💼 Tajriba: 0 yil")
    await query.message.reply_text(
        "7. 🏬 Oldingi ish joyingizda necha yil ishlagansiz?\nLavozimingizni ham yozishingiz mumkin."
    )
    return PREVIOUS_JOB


async def invalid_experience_choice(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> int:
    await reply_to_update(
        update,
        "Iltimos, tajriba bo'yicha Ha yoki Yo'q tugmasini tanlang.",
        reply_markup=EXPERIENCE_INLINE_KEYBOARD,
    )
    return EXPERIENCE_CHOICE


async def experience_years(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if not value:
        await update.message.reply_text("Iltimos, tajribangizni yozing. Masalan: 3 yil")
        return EXPERIENCE_YEARS

    if value.isdigit():
        value = f"{value} yil"

    context.user_data["experience"] = value
    if context.user_data.get("job_direction_key") in {"teacher", "assistant"}:
        await update.message.reply_text(
            "9. 📜 Sertifikatingiz bormi?",
            reply_markup=CERTIFICATE_CHOICE_KEYBOARD,
        )
        return CERTIFICATE_CHOICE

    await update.message.reply_text(
        "7. 🏬 Oldingi ish joyingizda necha yil ishlagansiz?\nLavozimingizni ham yozishingiz mumkin."
    )
    return PREVIOUS_JOB


async def certificate_choice_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    choice = query.data.split(":", 1)[1]
    await query.edit_message_reply_markup(reply_markup=None)
    if choice == "yes":
        context.user_data["certificate_status"] = "Bor"
        await query.message.reply_text("📜 Sertifikat rasmini yuboring.")
        return CERTIFICATE_PHOTO

    context.user_data["certificate_status"] = "Yo'q"
    context.user_data["certificate_image"] = {}
    await query.message.reply_text("📜 Sertifikat: Yo'q")
    await query.message.reply_text(
        "10. 🏬 Oldingi ish joyingizda necha yil ishlagansiz?\nLavozimingizni ham yozishingiz mumkin."
    )
    return PREVIOUS_JOB


async def invalid_certificate_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, sertifikat bor yoki yo'qligini tugma orqali tanlang.",
        reply_markup=CERTIFICATE_CHOICE_KEYBOARD,
    )
    return CERTIFICATE_CHOICE


async def certificate_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    image = get_image_info(update)
    if not image:
        await update.message.reply_text("Iltimos, sertifikat rasmini yuboring.")
        return CERTIFICATE_PHOTO

    context.user_data["certificate_status"] = "Bor"
    context.user_data["certificate_image"] = image
    await update.message.reply_text("✅ Sertifikat rasmi qabul qilindi.")
    if context.user_data.pop("return_to_review_after_certificate", None):
        await show_application_review(update.message, context)
        return REVIEW_APPLICATION

    await update.message.reply_text(
        "10. 🏬 Oldingi ish joyingizda necha yil ishlagansiz?\nLavozimingizni ham yozishingiz mumkin."
    )
    return PREVIOUS_JOB


async def invalid_certificate_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Iltimos, sertifikat rasmini yuboring.")
    return CERTIFICATE_PHOTO


async def previous_job(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Oldingi ish joyingiz haqida qisqacha yozing.")
        return PREVIOUS_JOB

    context.user_data["previous_job"] = value
    await update.message.reply_text(
        "8. ⚖️ Sudlanganmisiz?",
        reply_markup=CONVICTED_KEYBOARD,
    )
    return CONVICTED


async def convicted_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["convicted"] = YES_NO_LABELS[key]
    await query.edit_message_text(f"8. ⚖️ Sudlangan: {YES_NO_LABELS[key]}")
    await query.message.reply_text(
        "9. 👪 Oilaviy holatingiz qanday? Farzandlaringiz bormi?"
    )
    return FAMILY_STATUS


async def invalid_convicted(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, sudlanganlik bo'yicha Ha yoki Yo'q tugmasini tanlang.",
        reply_markup=CONVICTED_KEYBOARD,
    )
    return CONVICTED


async def family_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Oilaviy holatingizni qisqacha yozing.")
        return FAMILY_STATUS

    context.user_data["family_status"] = value
    await update.message.reply_text("10. 💰 Oxirgi ish joyingizda qancha maosh olgansiz?")
    return PREVIOUS_SALARY


async def previous_salary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Iltimos, oldingi maoshingizni yozing.")
        return PREVIOUS_SALARY

    context.user_data["previous_salary"] = value
    await update.message.reply_text("11. 💵 Qancha maoshga ishlashni xohlaysiz?")
    return EXPECTED_SALARY


async def expected_salary(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Iltimos, xohlagan maoshingizni yozing.")
        return EXPECTED_SALARY

    context.user_data["expected_salary"] = value
    await update.message.reply_text(
        "12. 📝 Microsoft Word dasturini bilish darajangizni tanlang.",
        reply_markup=WORD_LEVEL_KEYBOARD,
    )
    return WORD_LEVEL


async def word_level_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["word_level"] = LEVEL_LABELS[key]
    await query.edit_message_text(f"12. 📝 Word: {LEVEL_LABELS[key]}")
    await query.message.reply_text(
        "13. 📊 Microsoft Excel dasturini bilish darajangizni tanlang.",
        reply_markup=EXCEL_LEVEL_KEYBOARD,
    )
    return EXCEL_LEVEL


async def invalid_word_level(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, Word darajasini tugmalardan tanlang.",
        reply_markup=WORD_LEVEL_KEYBOARD,
    )
    return WORD_LEVEL


async def excel_level_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    key = query.data.split(":", 1)[1]
    context.user_data["excel_level"] = LEVEL_LABELS[key]
    await query.edit_message_text(f"13. 📊 Excel: {LEVEL_LABELS[key]}")
    await query.message.reply_text(
        "14. 🌐 Qaysi tillarni bilasiz va qay darajada?\nMasalan: O'zbekcha yaxshi, Ruscha o'rtacha"
    )
    return LANGUAGES


async def invalid_excel_level(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, Excel darajasini tugmalardan tanlang.",
        reply_markup=EXCEL_LEVEL_KEYBOARD,
    )
    return EXCEL_LEVEL


async def languages(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 2:
        await update.message.reply_text("Iltimos, biladigan tillaringizni yozing.")
        return LANGUAGES

    context.user_data["languages"] = value
    await update.message.reply_text(
        "15. ⏳ “Fariks O'quv markazi”da necha yil ishlash niyatingiz bor?"
    )
    return FARIKS_DURATION


async def fariks_duration(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 1:
        await update.message.reply_text("Iltimos, necha yil ishlash niyatingiz borligini yozing.")
        return FARIKS_DURATION

    context.user_data["fariks_duration"] = value
    await update.message.reply_text(
        "16. ❓ Nima uchun aynan “Fariks”da ishlashni xohlaysiz?"
    )
    return MOTIVATION


async def motivation(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if len(value) < 3:
        await update.message.reply_text("Iltimos, sababini qisqacha yozing.")
        return MOTIVATION

    context.user_data["motivation"] = value
    await update.message.reply_text("17. 📞 Telefon raqamingizni yozing. Masalan: +998901234567")
    return PHONE


async def phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    value = update.message.text.strip()
    if not phone_pattern.match(value):
        await update.message.reply_text(
            "Telefon raqamini to'g'ri formatda yuboring. Masalan: +998901234567"
        )
        return PHONE

    context.user_data["phone"] = value
    await show_application_review(update.message, context)
    return REVIEW_APPLICATION


def build_review_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data="review:confirm"),
                InlineKeyboardButton("✏️ Tahrirlash", callback_data="review:edit"),
            ],
            [InlineKeyboardButton("❌ Bekor qilish", callback_data="review:cancel")],
        ]
    )


def build_edit_fields_keyboard() -> InlineKeyboardMarkup:
    fields = [
        ("👤 Ism", "full_name"),
        ("🎂 Tug'ilgan sana", "birth_date"),
        ("📍 Manzil", "address"),
        ("🏢 Filial", "branch"),
        ("🎓 Ma'lumot", "education"),
        ("🚩 Yo'nalish", "job_direction"),
        ("📚 Mutaxassislik", "specialty"),
        ("📜 Sertifikat", "certificate_status"),
        ("💼 Tajriba", "experience"),
        ("🏬 Oldingi ish", "previous_job"),
        ("⚖️ Sudlangan", "convicted"),
        ("👪 Oilaviy holat", "family_status"),
        ("💰 Oldingi maosh", "previous_salary"),
        ("💵 Kutilgan maosh", "expected_salary"),
        ("📝 Word", "word_level"),
        ("📊 Excel", "excel_level"),
        ("🌐 Tillar", "languages"),
        ("⏳ Ishlash niyati", "fariks_duration"),
        ("❓ Nega Fariks", "motivation"),
        ("📞 Telefon", "phone"),
    ]
    rows = []
    for index in range(0, len(fields), 2):
        rows.append(
            [
                InlineKeyboardButton(label, callback_data=f"edit:{field}")
                for label, field in fields[index : index + 2]
            ]
        )
    rows.append([InlineKeyboardButton("⬅️ Orqaga", callback_data="edit:back")])
    return InlineKeyboardMarkup(rows)


def build_edit_value_keyboard(field: str) -> InlineKeyboardMarkup:
    if field == "education":
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("🎓 Oliy ma'lumotli", callback_data="editval:education:higher")],
                [InlineKeyboardButton("🏥 O'rta maxsus", callback_data="editval:education:secondary")],
            ]
        )

    if field == "convicted":
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ Ha", callback_data="editval:convicted:yes"),
                    InlineKeyboardButton("❌ Yo'q", callback_data="editval:convicted:no"),
                ]
            ]
        )

    if field == "job_direction":
        return InlineKeyboardMarkup(
            [
                [InlineKeyboardButton("🛠 Admin", callback_data="editval:job_direction:admin")],
                [InlineKeyboardButton("👨‍🏫 O'qituvchi", callback_data="editval:job_direction:teacher")],
                [InlineKeyboardButton("🤝 O'qituvchi yordamchi", callback_data="editval:job_direction:assistant")],
            ]
        )

    if field == "specialty":
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("📐 Matematika", callback_data="editval:specialty:math"),
                    InlineKeyboardButton("⚛️ Fizika", callback_data="editval:specialty:physics"),
                ],
                [
                    InlineKeyboardButton("🇷🇺 Rus tili", callback_data="editval:specialty:russian"),
                    InlineKeyboardButton("🇬🇧 Ingliz tili", callback_data="editval:specialty:english"),
                ],
            ]
        )

    if field == "certificate_status":
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("✅ Bor", callback_data="editval:certificate_status:yes"),
                    InlineKeyboardButton("❌ Yo'q", callback_data="editval:certificate_status:no"),
                ]
            ]
        )

    if field in {"word_level", "excel_level"}:
        prefix = "word_level" if field == "word_level" else "excel_level"
        return InlineKeyboardMarkup(
            [
                [
                    InlineKeyboardButton("❌ Bilmayman", callback_data=f"editval:{prefix}:unknown"),
                    InlineKeyboardButton("🔹 Bazaviy", callback_data=f"editval:{prefix}:basic"),
                ],
                [
                    InlineKeyboardButton("🔸 O'rtacha", callback_data=f"editval:{prefix}:medium"),
                    InlineKeyboardButton("✅ Yaxshi", callback_data=f"editval:{prefix}:good"),
                ],
            ]
        )

    return InlineKeyboardMarkup([])


def build_review_text(data: dict) -> str:
    text = (
        "📋 Arizangizni tekshirib chiqing:\n\n"
        f"👤 Ism-sharif: {data.get('full_name', '')}\n"
        f"🎂 Tug'ilgan sana: {data.get('birth_date', '')}\n"
        f"🔢 Yosh: {data.get('age', '')}\n"
        f"📍 Manzil: {data.get('address', '')}\n"
        f"🏢 Filial: {data.get('branch', '')}\n"
        f"🎓 Ma'lumot: {data.get('education', '')}\n"
        f"🚩 Yo'nalish: {data.get('job_direction', '')}\n"
    )
    if data.get("specialty"):
        text += f"📚 Mutaxassislik: {data.get('specialty', '')}\n"
    if data.get("certificate_status"):
        text += f"📜 Sertifikat: {data.get('certificate_status', '')}\n"

    text += (
        f"💼 Soha tajribasi: {data.get('experience', '')}\n"
        f"🏬 Oldingi ish joyi: {data.get('previous_job', '')}\n"
        f"⚖️ Sudlangan: {data.get('convicted', '')}\n"
        f"👪 Oilaviy holati: {data.get('family_status', '')}\n"
        f"💰 Oldingi maosh: {data.get('previous_salary', '')}\n"
        f"💵 Kutilayotgan maosh: {data.get('expected_salary', '')}\n"
        f"📝 Word: {data.get('word_level', '')}\n"
        f"📊 Excel: {data.get('excel_level', '')}\n"
        f"🌐 Tillar: {data.get('languages', '')}\n"
        f"⏳ Fariksda ishlash niyati: {data.get('fariks_duration', '')}\n"
        f"❓ Nega Fariks: {data.get('motivation', '')}\n"
        f"📞 Telefon: {data.get('phone', '')}"
    )
    return text


async def show_application_review(message, context: ContextTypes.DEFAULT_TYPE) -> None:
    await message.reply_text(
        build_review_text(context.user_data),
        reply_markup=build_review_keyboard(),
    )


async def review_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    action = query.data.split(":", 1)[1]
    if action == "confirm":
        await query.edit_message_reply_markup(reply_markup=None)
        if needs_certificate_image(context.user_data):
            context.user_data["return_to_review_after_certificate"] = True
            await query.message.reply_text("📜 Sertifikat rasmini yuboring.")
            return CERTIFICATE_PHOTO

        await query.message.reply_text("18. 📷 Oxirgi 1 oy ichida tushgan rasmingizni yuboring.")
        return RECENT_PHOTO

    if action == "edit":
        await query.edit_message_reply_markup(reply_markup=None)
        await query.message.reply_text(
            "Qaysi ma'lumotni tahrirlaysiz?",
            reply_markup=build_edit_fields_keyboard(),
        )
        return REVIEW_APPLICATION

    context.user_data.clear()
    await query.edit_message_text("Ariza to'ldirish bekor qilindi.")
    await query.message.reply_text(
        "Qayta boshlash uchun /start bosing.",
        reply_markup=main_keyboard_for(query.from_user.id),
    )
    return ConversationHandler.END


def needs_certificate_image(data: dict) -> bool:
    is_teacher_direction = data.get("job_direction_key") in {"teacher", "assistant"} or data.get(
        "job_direction"
    ) in {"O'qituvchi", "O'qituvchi yordamchi"}
    return (
        is_teacher_direction
        and data.get("certificate_status") == "Bor"
        and not data.get("certificate_image")
    )


async def edit_field_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    field = query.data.split(":", 1)[1]
    if field == "back":
        await query.edit_message_reply_markup(reply_markup=None)
        await show_application_review(query.message, context)
        return REVIEW_APPLICATION

    if field in {
        "education",
        "convicted",
        "word_level",
        "excel_level",
        "job_direction",
        "specialty",
        "certificate_status",
    }:
        await query.message.reply_text(
            edit_field_prompt(field),
            reply_markup=build_edit_value_keyboard(field),
        )
        return REVIEW_APPLICATION

    context.user_data["editing_field"] = field
    await query.message.reply_text(edit_field_prompt(field))
    return EDIT_FIELD


async def edit_value_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    _, field, key = query.data.split(":")
    if field == "education":
        context.user_data[field] = EDUCATION_LABELS[key]
    elif field == "job_direction":
        context.user_data[field] = JOB_DIRECTION_LABELS[key]
        context.user_data["job_direction_key"] = key
        if key == "admin":
            context.user_data["specialty"] = ""
            context.user_data["certificate_status"] = ""
            context.user_data["certificate_image"] = {}
    elif field == "specialty":
        context.user_data[field] = SPECIALTY_LABELS[key]
    elif field == "certificate_status":
        context.user_data[field] = "Bor" if key == "yes" else "Yo'q"
        if key == "no":
            context.user_data["certificate_image"] = {}
    elif field == "convicted":
        context.user_data[field] = YES_NO_LABELS[key]
    elif field in {"word_level", "excel_level"}:
        context.user_data[field] = LEVEL_LABELS[key]

    await query.edit_message_reply_markup(reply_markup=None)
    await query.message.reply_text("✅ Ma'lumot yangilandi.")
    await show_application_review(query.message, context)
    return REVIEW_APPLICATION


async def edit_field_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    field = context.user_data.get("editing_field")
    if not field:
        await show_application_review(update.message, context)
        return REVIEW_APPLICATION

    value = update.message.text.strip()
    error = validate_edit_value(field, value)
    if error:
        await update.message.reply_text(error)
        return EDIT_FIELD

    if field == "birth_date":
        context.user_data["birth_date"] = value
        context.user_data["age"] = str(calculate_age(value))
    elif field == "experience" and value.isdigit():
        context.user_data[field] = f"{value} yil"
    else:
        context.user_data[field] = value

    context.user_data.pop("editing_field", None)
    await update.message.reply_text("✅ Ma'lumot yangilandi.")
    await show_application_review(update.message, context)
    return REVIEW_APPLICATION


async def invalid_review_input(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await reply_to_update(
        update,
        "Iltimos, pastdagi tugmalardan birini tanlang.",
        reply_markup=build_review_keyboard(),
    )
    return REVIEW_APPLICATION


def edit_field_prompt(field: str) -> str:
    prompts = {
        "full_name": "Yangi ism-sharifni yozing.",
        "birth_date": "Yangi tug'ilgan sanani yozing. Masalan: 29.10.2000",
        "address": "Yangi yashash manzilini yozing.",
        "branch": "Yangi filial shahar yoki tuman nomini yozing.",
        "education": "Yangi ma'lumot darajasini tanlang.",
        "job_direction": "Yangi yo'nalishni tanlang.",
        "specialty": "Yangi mutaxassislikni tanlang.",
        "certificate_status": "Sertifikat bor yoki yo'qligini tanlang.",
        "experience": "Yangi tajribani yozing. Masalan: 3 yil",
        "previous_job": "Oldingi ish joyi haqida yangi ma'lumot yozing.",
        "convicted": "Sudlanganlik holatini tanlang.",
        "family_status": "Oilaviy holatni qayta yozing.",
        "previous_salary": "Oldingi maoshni qayta yozing.",
        "expected_salary": "Kutilayotgan maoshni qayta yozing.",
        "word_level": "Word darajasini tanlang.",
        "excel_level": "Excel darajasini tanlang.",
        "languages": "Tillarni qayta yozing.",
        "fariks_duration": "Fariksda necha yil ishlash niyatingiz borligini yozing.",
        "motivation": "Nima uchun Fariksda ishlashni xohlashingizni qayta yozing.",
        "phone": "Yangi telefon raqamni yozing. Masalan: +998901234567",
    }
    return prompts.get(field, "Yangi qiymatni yozing.")


def validate_edit_value(field: str, value: str) -> str | None:
    if field == "full_name" and len(value) < 3:
        return "Iltimos, ism-sharifingizni to'liq yozing."
    if field == "birth_date" and not is_valid_birth_date(value):
        return "Sanani to'g'ri yozing. Masalan: 29.10.2000"
    if field == "address" and len(value) < 5:
        return "Iltimos, manzilingizni to'liqroq yozing."
    if field == "branch" and len(value) < 2:
        return "Iltimos, filial shahar yoki tuman nomini yozing."
    if field == "phone" and not phone_pattern.match(value):
        return "Telefon raqamini to'g'ri formatda yuboring. Masalan: +998901234567"
    if len(value) < 1:
        return "Iltimos, qiymatni yozing."
    return None


async def recent_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    image = get_image_info(update)
    if not image:
        await update.message.reply_text("Iltimos, rasm yuboring.")
        return RECENT_PHOTO

    context.user_data["recent_photo"] = image
    return await finish_application(update, context)


def get_image_info(update: Update) -> dict | None:
    message = update.message
    if message.photo:
        return {
            "type": "photo",
            "file_id": message.photo[-1].file_id,
            "mime_type": "image/jpeg",
        }

    if message.document and message.document.mime_type:
        if message.document.mime_type.startswith("image/"):
            return {
                "type": "document",
                "file_id": message.document.file_id,
                "mime_type": message.document.mime_type,
                "file_name": message.document.file_name,
            }

    return None


async def invalid_recent_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text("Iltimos, oxirgi 1 oy ichida tushgan rasmingizni yuboring.")
    return RECENT_PHOTO


async def finish_application(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    upsert_user(user)

    required_fields = [
        "full_name",
        "birth_date",
        "age",
        "address",
        "branch",
        "education",
        "job_direction",
        "experience",
        "previous_job",
        "convicted",
        "family_status",
        "previous_salary",
        "expected_salary",
        "word_level",
        "excel_level",
        "languages",
        "fariks_duration",
        "motivation",
        "phone",
        "recent_photo",
    ]
    is_teacher_direction = context.user_data.get("job_direction_key") in {
        "teacher",
        "assistant",
    } or context.user_data.get("job_direction") in {"O'qituvchi", "O'qituvchi yordamchi"}
    if is_teacher_direction:
        required_fields.extend(["specialty", "certificate_status"])
        if context.user_data.get("certificate_status") == "Bor":
            required_fields.append("certificate_image")

    if any(not context.user_data.get(field) for field in required_fields):
        await reply_to_update(update, "Ariza ma'lumotlari to'liq emas. /start orqali qayta boshlang.")
        context.user_data.clear()
        return ConversationHandler.END

    application_id = create_application(user, context.user_data)
    application = row_to_application(get_application(application_id))

    try:
        await send_application_to_admin(context, application)
    except Exception as error:
        logger.error("Adminlarga yuborishda xatolik: %s", format_runtime_error(error))
        delete_application(application_id)
        await reply_to_update(
            update,
            "Kechirasiz, arizani adminlarga yuborishda xatolik yuz berdi. "
            "Iltimos, keyinroq qayta urinib ko'ring.",
            reply_markup=main_keyboard_for(user.id),
        )
        context.user_data.clear()
        return ConversationHandler.END

    await reply_to_update(
        update,
        "✅ Arizangiz adminlarga yuborildi. Javobni kuting.",
        reply_markup=main_keyboard_for(user.id),
    )
    context.user_data.clear()
    return ConversationHandler.END


async def reply_to_update(
    update: Update,
    text: str,
    reply_markup=None,
) -> None:
    if update.message:
        await update.message.reply_text(text, reply_markup=reply_markup)
        return

    if update.callback_query and update.callback_query.message:
        await update.callback_query.message.reply_text(text, reply_markup=reply_markup)


async def send_application_to_admin(
    context: ContextTypes.DEFAULT_TYPE,
    application: dict,
) -> None:
    chat_ids = get_admin_delivery_chat_ids()
    if not chat_ids:
        raise RuntimeError("ADMIN_CHAT_ID .env faylida ko'rsatilmagan.")

    admin_messages = []
    for chat_id in chat_ids:
        try:
            sent_messages = await send_application_to_chat(
                context=context,
                chat_id=chat_id,
                application=application,
                reply_markup=build_decision_keyboard(application["id"]),
            )
            admin_messages.extend(
                {
                    "chat_id": str(message.chat_id),
                    "message_id": message.message_id,
                }
                for message in sent_messages
            )
        except Exception as error:
            logger.warning(
                "Adminga yuborilmadi (%s): %s",
                chat_id,
                format_runtime_error(error),
            )

    if not admin_messages:
        raise RuntimeError("Arizani birorta adminga yuborib bo'lmadi.")

    save_application_admin_messages(application["id"], admin_messages)


async def send_application_to_chat(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: str,
    application: dict,
    reply_markup: InlineKeyboardMarkup | None = None,
) -> list:
    caption = build_application_caption(application)
    image = application.get("recent_photo") or {}
    certificate_image = application.get("certificate_image") or {}
    sent_messages = []

    if image and certificate_image:
        album_caption = caption if len(caption) <= 1024 else limit_telegram_text(caption, 1024)
        album_messages = await send_image_album(
            context=context,
            chat_id=chat_id,
            images=[image, certificate_image],
            caption=album_caption,
        )
        sent_messages.extend(album_messages)

        if len(caption) > 1024:
            text_message = await context.bot.send_message(
                chat_id=chat_id,
                text=limit_telegram_text(caption),
                reply_markup=reply_markup,
            )
            sent_messages.append(text_message)
        elif reply_markup:
            control_message = await context.bot.send_message(
                chat_id=chat_id,
                text=f"Ariza #{application['id']} bo'yicha qaror tanlang:",
                reply_markup=reply_markup,
            )
            sent_messages.append(control_message)

        return sent_messages

    if not image:
        message = await context.bot.send_message(
            chat_id=chat_id,
            text=limit_telegram_text(caption),
            reply_markup=reply_markup,
        )
        sent_messages.append(message)
    elif len(caption) <= 1024:
        message = await send_single_image(
            context=context,
            chat_id=chat_id,
            image=image,
            caption=caption,
            reply_markup=reply_markup,
        )
        sent_messages.append(message)
    else:
        image_message = await send_single_image(context=context, chat_id=chat_id, image=image, caption=None)
        text_message = await context.bot.send_message(
            chat_id=chat_id,
            text=limit_telegram_text(caption),
            reply_markup=reply_markup,
        )
        sent_messages.extend([image_message, text_message])

    if certificate_image:
        certificate_message = await send_single_image(
            context=context,
            chat_id=chat_id,
            image=certificate_image,
            caption="📜 Sertifikat",
        )
        sent_messages.append(certificate_message)

    return sent_messages


async def send_image_album(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: str,
    images: list[dict],
    caption: str,
) -> list:
    try:
        return await context.bot.send_media_group(
            chat_id=chat_id,
            media=[
                build_album_media(image, caption if index == 0 else None)
                for index, image in enumerate(images)
            ],
        )
    except Exception as error:
        logger.warning("Album yuborishda xatolik: %s", format_runtime_error(error))
        sent_messages = []
        for index, image in enumerate(images):
            sent_messages.append(
                await send_single_image(
                    context=context,
                    chat_id=chat_id,
                    image=image,
                    caption=caption if index == 0 else None,
                )
            )
        return sent_messages


def build_album_media(image: dict, caption: str | None):
    if image.get("type") == "photo":
        return InputMediaPhoto(media=image["file_id"], caption=caption)
    return InputMediaDocument(media=image["file_id"], caption=caption)


async def send_single_image(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: str,
    image: dict,
    caption: str | None,
    reply_markup: InlineKeyboardMarkup | None = None,
):
    if image.get("type") == "photo":
        return await context.bot.send_photo(
            chat_id=chat_id,
            photo=image["file_id"],
            caption=caption,
            reply_markup=reply_markup,
        )

    return await context.bot.send_document(
        chat_id=chat_id,
        document=image["file_id"],
        caption=caption,
        reply_markup=reply_markup,
    )


def build_application_caption(application: dict) -> str:
    text = (
        "🆕 Yangi nomzod\n\n"
        f"👤 Ism-sharif: {application['full_name']}\n"
        f"🎂 Tug'ilgan sana: {application['birth_date']}\n"
        f"🔢 Yosh: {application['age']}\n"
        f"📍 Manzil: {application['address']}\n"
        f"🏢 Filial: {application['branch']}\n"
        f"🎓 Ma'lumot: {application['education']}\n"
        f"🚩 Yo'nalish: {application['job_direction']}\n"
    )
    if application.get("specialty"):
        text += f"📚 Mutaxassislik: {application['specialty']}\n"
    if application.get("certificate_status"):
        text += f"📜 Sertifikat: {application['certificate_status']}\n"

    text += (
        f"💼 Soha tajribasi: {application['experience']}\n"
        f"🏬 Oldingi ish joyi: {application['previous_job']}\n"
        f"⚖️ Sudlangan: {application['convicted']}\n"
        f"👪 Oilaviy holati: {application['family_status']}\n"
        f"💰 Oldingi maosh: {application['previous_salary']}\n"
        f"💵 Kutilayotgan maosh: {application['expected_salary']}\n"
        f"📝 Word: {application['word_level']}\n"
        f"📊 Excel: {application['excel_level']}\n"
        f"🌐 Tillar: {application['languages']}\n"
        f"⏳ Fariksda ishlash niyati: {application['fariks_duration']}\n"
        f"❓ Nega Fariks: {application['motivation']}\n"
        f"📞 Telefon: {application['phone']}"
    )
    if application.get("reject_reason"):
        text += f"\n📝 Rad sababi: {application['reject_reason']}"
    return text


def limit_telegram_text(text: str, limit: int = 3900) -> str:
    if len(text) <= limit:
        return text
    return text[: limit - 20].rstrip() + "\n\n... qisqartirildi"


def build_decision_keyboard(application_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"app:approve:{application_id}"),
                InlineKeyboardButton("❌ Rad etish", callback_data=f"app:reject:{application_id}"),
            ]
        ]
    )


async def handle_application_decision(
    update: Update,
    context: ContextTypes.DEFAULT_TYPE,
) -> None:
    query = update.callback_query
    admin_user = query.from_user
    upsert_user(admin_user)

    if not is_admin(admin_user.id):
        await query.answer("Bu amal faqat adminlar uchun.", show_alert=True)
        return

    _, action, application_id_text = query.data.split(":")
    application_id = int(application_id_text)
    row = get_application(application_id)

    if not row:
        await query.answer("Ariza topilmadi.", show_alert=True)
        return

    application = row_to_application(row)
    if application["status"] != "pending":
        await query.answer(
            f"Bu ariza allaqachon {status_label(application['status']).lower()}.",
            show_alert=True,
        )
        await remove_decision_buttons(query)
        return

    if action == "approve":
        if not PUBLISH_CHAT_ID:
            await query.answer("PUBLISH_CHAT_ID sozlanmagan.", show_alert=True)
            return

        locked = try_update_application_status(
            application_id,
            expected_status="pending",
            status="processing",
            admin_id=admin_user.id,
        )
        if not locked:
            fresh_application = row_to_application(get_application(application_id))
            await query.answer(
                f"Bu ariza allaqachon {status_label(fresh_application['status']).lower()}.",
                show_alert=True,
            )
            await remove_decision_buttons(query)
            return

        try:
            await send_application_to_chat(
                context=context,
                chat_id=PUBLISH_CHAT_ID,
                application=application,
            )
        except Exception as error:
            set_application_status(application_id, "pending", admin_user.id)
            logger.error("Kanalga yuborishda xatolik: %s", format_runtime_error(error))
            await query.answer("Kanalga yuborishda xatolik bo'ldi.", show_alert=True)
            return

        set_application_status(application_id, "approved", admin_user.id)
        approved_application = row_to_application(get_application(application_id))
        await notify_user(
            context,
            application["user_id"],
            "Arizangiz tasdiqlandi. Siz bilan yaqinda bog'lanamiz. ✅",
        )
        await remove_decision_buttons_from_admin_messages(context, application_id)
        await notify_admins_about_decision(context, approved_application, admin_user, "approved")
        await query.answer("Ariza tasdiqlandi.")
        return

    admin_reject_targets[admin_user.id] = application_id
    await remove_decision_buttons(query)
    await query.answer("Rad etish sababini yozing.")
    await query.message.reply_text(
        f"❌ Ariza #{application_id} uchun rad etish sababini yozing.\n"
        "Masalan: Tajriba yetarli emas.\n\nBekor qilish uchun /cancel.",
    )


async def remove_decision_buttons(query) -> None:
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception as error:
        logger.warning("Inline tugmalarni o'chirishda xatolik: %s", format_runtime_error(error))


async def remove_decision_buttons_from_admin_messages(
    context: ContextTypes.DEFAULT_TYPE,
    application_id: int,
) -> None:
    for message in get_application_admin_messages(application_id):
        try:
            await context.bot.edit_message_reply_markup(
                chat_id=message["chat_id"],
                message_id=message["message_id"],
                reply_markup=None,
            )
        except Exception:
            pass


async def notify_admins_about_decision(
    context: ContextTypes.DEFAULT_TYPE,
    application: dict,
    admin_user,
    status: str,
    reason: str = "",
) -> None:
    admin_name = f"@{admin_user.username}" if admin_user.username else str(admin_user.id)
    if status == "approved":
        text = (
            f"✅ Ariza #{application['id']} tasdiqlandi.\n"
            f"👮 Qaror qilgan admin: {admin_name}\n"
            f"👤 Nomzod: {application['full_name']}"
        )
    else:
        text = (
            f"❌ Ariza #{application['id']} rad etildi.\n"
            f"👮 Qaror qilgan admin: {admin_name}\n"
            f"👤 Nomzod: {application['full_name']}\n"
            f"📝 Sabab: {reason}"
        )

    for chat_id in get_admin_delivery_chat_ids():
        try:
            await context.bot.send_message(chat_id=chat_id, text=text)
        except Exception:
            pass


async def notify_user(context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str) -> None:
    try:
        await context.bot.send_message(chat_id=user_id, text=text)
    except Exception as error:
        logger.warning("Foydalanuvchiga xabar bormadi (%s): %s", user_id, format_runtime_error(error))


def sync_excel_file(
    rows: list[sqlite3.Row] | None = None,
    image_paths: dict[int, Path] | None = None,
) -> Path:
    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Arizalar"
    sheet.sheet_view.showGridLines = False
    image_paths = image_paths or {}

    headers = [
        "No",
        "Ism-sharif",
        "Tug'ilgan sana",
        "Yosh",
        "Telefon",
        "Manzil",
        "Filial",
        "Ma'lumot",
        "Yo'nalish",
        "Mutaxassislik",
        "Soha tajribasi",
        "Sertifikat",
        "Oldingi ish joyi",
        "Sudlangan",
        "Oilaviy holati",
        "Oldingi maosh",
        "Kutilayotgan maosh",
        "Word",
        "Excel",
        "Tillar",
        "Fariksda ishlash niyati",
        "Nega Fariks",
        "Rasm",
    ]
    last_column = get_column_letter(len(headers))
    sheet.append(headers)
    image_column = headers.index("Rasm") + 1
    application_rows = rows if rows is not None else get_approved_application_rows()
    header_row = 1

    for index, row in enumerate(application_rows, start=1):
        application = row_to_application(row)
        photo_exists = bool((application.get("recent_photo") or {}).get("file_id"))
        sheet.append(
            [
                index,
                application["full_name"],
                application["birth_date"],
                application["age"],
                application["phone"],
                application["address"],
                application["branch"],
                application["education"],
                application["job_direction"],
                application["specialty"],
                application["experience"],
                application["certificate_status"],
                application["previous_job"],
                application["convicted"],
                application["family_status"],
                application["previous_salary"],
                application["expected_salary"],
                application["word_level"],
                application["excel_level"],
                application["languages"],
                application["fariks_duration"],
                application["motivation"],
                "" if image_paths.get(application["id"]) else ("Rasm bor" if photo_exists else "Rasm yo'q"),
            ]
        )
        row_number = sheet.max_row
        image_path = image_paths.get(application["id"])
        if image_path:
            excel_image = ExcelImage(str(image_path))
            excel_image.width = 105
            excel_image.height = 105
            sheet.add_image(excel_image, f"{get_column_letter(image_column)}{row_number}")
            sheet.row_dimensions[row_number].height = 82
            sheet.cell(row=row_number, column=image_column).alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    apply_excel_design(sheet, headers, header_row, image_column)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{sheet.max_row}"

    workbook.save(EXCEL_FILE)
    return EXCEL_FILE


def apply_excel_design(sheet, headers: list[str], header_row: int, image_column: int) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    row_fill = PatternFill("solid", fgColor="FFFFFF")

    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    sheet.row_dimensions[header_row].height = 24

    for row_number in range(header_row + 1, sheet.max_row + 1):
        for cell in sheet[row_number]:
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    width_by_header = {
        "No": 6,
        "Ism-sharif": 24,
        "Tug'ilgan sana": 14,
        "Yosh": 7,
        "Telefon": 18,
        "Manzil": 28,
        "Filial": 20,
        "Ma'lumot": 18,
        "Yo'nalish": 20,
        "Mutaxassislik": 18,
        "Soha tajribasi": 16,
        "Sertifikat": 14,
        "Oldingi ish joyi": 26,
        "Sudlangan": 12,
        "Oilaviy holati": 26,
        "Oldingi maosh": 16,
        "Kutilayotgan maosh": 18,
        "Word": 12,
        "Excel": 12,
        "Tillar": 18,
        "Fariksda ishlash niyati": 18,
        "Nega Fariks": 28,
        "Rasm": 18,
    }
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width_by_header.get(header, 18)
    sheet.column_dimensions[get_column_letter(image_column)].width = 18


async def sync_excel_file_with_images(context: ContextTypes.DEFAULT_TYPE) -> Path:
    rows = get_approved_application_rows()
    with TemporaryDirectory() as temp_dir:
        image_paths = await download_excel_images(context, rows, Path(temp_dir))
        return sync_excel_file(rows=rows, image_paths=image_paths)


async def download_excel_images(
    context: ContextTypes.DEFAULT_TYPE,
    rows: list[sqlite3.Row],
    temp_dir: Path,
) -> dict[int, Path]:
    image_paths: dict[int, Path] = {}

    for row in rows:
        application = row_to_application(row)
        photo = application.get("recent_photo") or {}
        file_id = photo.get("file_id")
        if not file_id:
            continue

        original_path = temp_dir / f"app_{application['id']}_original"
        thumbnail_path = temp_dir / f"app_{application['id']}.jpg"
        try:
            telegram_file = await context.bot.get_file(file_id)
            await telegram_file.download_to_drive(custom_path=original_path)
            resize_image_for_excel(original_path, thumbnail_path)
            image_paths[application["id"]] = thumbnail_path
        except Exception as error:
            logger.warning(
                "Excel uchun rasm yuklanmadi (#%s): %s",
                application["id"],
                format_runtime_error(error),
            )

    return image_paths


def resize_image_for_excel(source_path: Path, target_path: Path) -> None:
    with PILImage.open(source_path) as image:
        image = ImageOps.exif_transpose(image)
        if image.mode not in {"RGB", "L"}:
            image = image.convert("RGB")
        image = ImageOps.fit(image, (180, 180), method=PILImage.Resampling.LANCZOS)
        image.save(target_path, format="JPEG", quality=90)


def safe_sync_excel_file() -> Path | None:
    try:
        return sync_excel_file()
    except Exception as error:
        logger.warning("Excel yangilashda xatolik: %s", format_runtime_error(error))
        return None


def status_label(status: str) -> str:
    return {
        "pending": "Kutilmoqda",
        "processing": "Ko'rib chiqilmoqda",
        "approved": "Tasdiqlangan",
        "rejected": "Rad etilgan",
    }.get(status, status)


def format_iso_datetime(value: str) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).astimezone().strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return value


async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    upsert_user(update.effective_user)
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Bu bo'lim faqat adminlar uchun.")
        return

    await update.message.reply_text(
        build_admin_panel_text(),
        reply_markup=build_admin_panel_keyboard(),
    )


async def my_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    upsert_user(update.effective_user)
    username = f"@{update.effective_user.username}" if update.effective_user.username else "username yo'q"
    await update.message.reply_text(
        f"Sizning Telegram ID: {update.effective_user.id}\nUsername: {username}"
    )


async def admin_panel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu bo'lim faqat adminlar uchun.", show_alert=True)
        return

    await query.answer()
    await query.edit_message_text(
        build_admin_panel_text(),
        reply_markup=build_admin_panel_keyboard(),
    )


async def admin_excel_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu bo'lim faqat adminlar uchun.", show_alert=True)
        return

    await query.answer("Excel tayyorlanyapti...")
    try:
        excel_path = await sync_excel_file_with_images(context)
    except Exception as error:
        logger.error("Excel yuborishda xatolik: %s", format_runtime_error(error))
        await query.message.reply_text(
            "Excel fayl tayyorlashda xatolik bo'ldi. Fayl ochiq bo'lsa yopib qayta urinib ko'ring."
        )
        return

    with excel_path.open("rb") as file:
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=file,
            filename=excel_path.name,
            caption="📊 Arizalar Excel fayli",
        )


def build_admin_panel_text() -> str:
    stats = get_stats()
    return (
        "🛠 Admin panel\n\n"
        f"👥 Foydalanuvchilar soni: {stats['users']}\n"
        f"👮 Adminlar soni: {stats['admins']}\n"
        f"📨 Kutilayotgan arizalar: {stats['pending']}\n"
        f"✅ Tasdiqlangan: {stats['approved']}\n"
        f"❌ Rad etilgan: {stats['rejected']}"
    )


def build_admin_panel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [
                InlineKeyboardButton("➕ Admin qo'shish", callback_data="admin:add"),
                InlineKeyboardButton("➖ Admin o'chirish", callback_data="admin:remove"),
            ],
            [
                InlineKeyboardButton("🟡 Kutilmoqda", callback_data="admin:list:pending"),
                InlineKeyboardButton("✅ Tasdiqlangan", callback_data="admin:list:approved"),
            ],
            [
                InlineKeyboardButton("❌ Rad etilgan", callback_data="admin:list:rejected"),
                InlineKeyboardButton("🔎 Qidirish", callback_data="admin:search"),
            ],
            [
                InlineKeyboardButton("🏢 Filial filter", callback_data="admin:filter_branch"),
                InlineKeyboardButton("📊 Excel", callback_data="admin:excel"),
            ],
            [InlineKeyboardButton("🔄 Yangilash", callback_data="admin:refresh")],
        ]
    )


async def admin_status_list_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu bo'lim faqat adminlar uchun.", show_alert=True)
        return

    status = query.data.split(":")[-1]
    rows = get_application_rows_by_status(status)
    await query.answer()
    await query.message.reply_text(
        build_application_list_text(f"{status_label(status)} arizalar", rows),
        reply_markup=build_application_list_keyboard(rows),
    )


async def admin_view_application_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu bo'lim faqat adminlar uchun.", show_alert=True)
        return

    application_id = int(query.data.split(":")[-1])
    row = get_application(application_id)
    if not row:
        await query.answer("Ariza topilmadi.", show_alert=True)
        return

    application = row_to_application(row)
    reply_markup = None
    if application["status"] == "pending":
        reply_markup = build_decision_keyboard(application["id"])

    await query.answer()
    await send_application_to_chat(
        context=context,
        chat_id=query.message.chat_id,
        application=application,
        reply_markup=reply_markup,
    )


async def admin_search_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu amal faqat adminlar uchun.", show_alert=True)
        return

    await query.answer()
    admin_pending_actions[query.from_user.id] = SEARCH_APPLICATION_TARGET
    await query.message.reply_text(
        "🔎 Qidirish uchun ism, telefon raqam yoki filial nomini yozing.\n"
        "Bekor qilish uchun /cancel.",
    )


async def admin_filter_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu amal faqat adminlar uchun.", show_alert=True)
        return

    await query.answer()
    admin_pending_actions[query.from_user.id] = FILTER_BRANCH_TARGET
    await query.message.reply_text(
        "🏢 Qaysi filial bo'yicha filter qilamiz? Shahar yoki tuman nomini yozing.\n"
        "Bekor qilish uchun /cancel.",
    )


async def admin_search_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query_text = update.message.text.strip()
    if len(query_text) < 2:
        await update.message.reply_text("Qidirish uchun kamida 2 ta belgi yozing.")
        return

    rows = search_application_rows(query_text)
    admin_pending_actions.pop(update.effective_user.id, None)
    await update.message.reply_text(
        build_application_list_text(f"Qidiruv: {query_text}", rows),
        reply_markup=build_application_list_keyboard(rows),
    )


async def admin_filter_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    branch_text = update.message.text.strip()
    if len(branch_text) < 2:
        await update.message.reply_text("Filial uchun kamida 2 ta belgi yozing.")
        return

    rows = get_application_rows_by_branch(branch_text)
    admin_pending_actions.pop(update.effective_user.id, None)
    await update.message.reply_text(
        build_application_list_text(f"Filial: {branch_text}", rows),
        reply_markup=build_application_list_keyboard(rows),
    )


def build_application_list_text(title: str, rows: list[sqlite3.Row]) -> str:
    if not rows:
        return f"📭 {title}\n\nAriza topilmadi."

    lines = [f"📋 {title}", ""]
    for row in rows:
        application = row_to_application(row)
        lines.append(
            f"#{application['id']} | {status_label(application['status'])} | "
            f"{application['full_name']} | {application['branch']} | {application['phone']}"
        )
    lines.append("")
    lines.append("Ko'rish uchun pastdagi tugmalardan birini bosing.")
    return "\n".join(lines)


def build_application_list_keyboard(rows: list[sqlite3.Row]) -> InlineKeyboardMarkup | None:
    if not rows:
        return None

    buttons = []
    for row in rows[:10]:
        application = row_to_application(row)
        name = application["full_name"][:22]
        buttons.append(
            [InlineKeyboardButton(f"👁 #{application['id']} {name}", callback_data=f"admin:view:{application['id']}")]
        )
    buttons.append([InlineKeyboardButton("🛠 Admin panel", callback_data="admin:refresh")])
    return InlineKeyboardMarkup(buttons)


async def admin_add_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu amal faqat adminlar uchun.", show_alert=True)
        return ConversationHandler.END

    await query.answer()
    admin_pending_actions[query.from_user.id] = ADD_ADMIN_TARGET
    await query.message.reply_text(
        "Yangi adminning Telegram ID raqamini yoki @username yuboring.\n"
        "Bekor qilish uchun /cancel.",
    )
    return ConversationHandler.END


async def admin_add_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Bu amal faqat adminlar uchun.")
        admin_pending_actions.pop(update.effective_user.id, None)
        return

    user_id, username, error = resolve_admin_target(update.message.text)
    if error:
        await update.message.reply_text(error)
        return

    add_admin(user_id, username, update.effective_user.id)
    admin_pending_actions.pop(update.effective_user.id, None)
    username_text = f"@{username}" if username else str(user_id)
    await update.message.reply_text(
        f"✅ {username_text} admin qilindi.",
        reply_markup=ADMIN_MAIN_KEYBOARD,
    )


async def admin_remove_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    upsert_user(query.from_user)

    if not is_admin(query.from_user.id):
        await query.answer("Bu amal faqat adminlar uchun.", show_alert=True)
        return

    await query.answer()
    admin_pending_actions[query.from_user.id] = REMOVE_ADMIN_TARGET
    await query.message.reply_text(
        "O'chiriladigan adminning Telegram ID raqamini yoki @username yuboring.\n"
        "Bekor qilish uchun /cancel.",
    )


async def admin_remove_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("Bu amal faqat adminlar uchun.")
        admin_pending_actions.pop(update.effective_user.id, None)
        return

    user_id, username, error = resolve_admin_target(update.message.text)
    if error:
        await update.message.reply_text(error)
        return

    if user_id in ADMIN_IDS:
        admin_pending_actions.pop(update.effective_user.id, None)
        await update.message.reply_text(
            "Bu admin .env ichidagi ADMIN_IDS orqali berilgan. Uni paneldan o'chirib bo'lmaydi.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )
        return

    removed = remove_admin(user_id)
    admin_pending_actions.pop(update.effective_user.id, None)
    username_text = f"@{username}" if username else str(user_id)
    if removed:
        await update.message.reply_text(
            f"✅ {username_text} adminlikdan olib tashlandi.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )
    else:
        await update.message.reply_text(
            "Bu foydalanuvchi adminlar ro'yxatida topilmadi.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )


async def admin_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    admin_pending_actions.pop(update.effective_user.id, None)
    admin_reject_targets.pop(update.effective_user.id, None)
    await update.message.reply_text("Amal bekor qilindi.", reply_markup=ADMIN_MAIN_KEYBOARD)
    return ConversationHandler.END


async def admin_pending_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.effective_user.id in admin_reject_targets:
        await admin_reject_reason_receive(update, context)
        return

    action = admin_pending_actions.get(update.effective_user.id)
    if action == ADD_ADMIN_TARGET:
        await admin_add_receive(update, context)
        return
    if action == REMOVE_ADMIN_TARGET:
        await admin_remove_receive(update, context)
        return
    if action == SEARCH_APPLICATION_TARGET:
        await admin_search_receive(update, context)
        return
    if action == FILTER_BRANCH_TARGET:
        await admin_filter_receive(update, context)


async def admin_reject_reason_receive(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_admin(update.effective_user.id):
        admin_reject_targets.pop(update.effective_user.id, None)
        await update.message.reply_text("Bu amal faqat adminlar uchun.")
        return

    application_id = admin_reject_targets.get(update.effective_user.id)
    reason = update.message.text.strip()
    if len(reason) < 3:
        await update.message.reply_text("Rad etish sababini to'liqroq yozing.")
        return

    row = get_application(application_id)
    if not row:
        admin_reject_targets.pop(update.effective_user.id, None)
        await update.message.reply_text("Ariza topilmadi.", reply_markup=ADMIN_MAIN_KEYBOARD)
        return

    application = row_to_application(row)
    if application["status"] != "pending":
        admin_reject_targets.pop(update.effective_user.id, None)
        await update.message.reply_text(
            f"Bu ariza allaqachon {status_label(application['status']).lower()}.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )
        return

    decided = try_update_application_status(
        application_id,
        expected_status="pending",
        status="rejected",
        admin_id=update.effective_user.id,
        reject_reason=reason,
    )
    admin_reject_targets.pop(update.effective_user.id, None)
    if not decided:
        fresh_row = get_application(application_id)
        fresh_application = row_to_application(fresh_row) if fresh_row else application
        await update.message.reply_text(
            f"Bu ariza allaqachon {status_label(fresh_application['status']).lower()}.",
            reply_markup=ADMIN_MAIN_KEYBOARD,
        )
        return

    rejected_application = row_to_application(get_application(application_id))
    await notify_user(
        context,
        application["user_id"],
        f"Arizangiz rad etildi. ❌\nSabab: {reason}",
    )
    await remove_decision_buttons_from_admin_messages(context, application_id)
    await notify_admins_about_decision(
        context,
        rejected_application,
        update.effective_user,
        "rejected",
        reason,
    )
    await update.message.reply_text(
        f"❌ Ariza #{application_id} rad etildi.\n📝 Sabab: {reason}",
        reply_markup=ADMIN_MAIN_KEYBOARD,
    )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text(
        "Ariza to'ldirish bekor qilindi.",
        reply_markup=main_keyboard_for(update.effective_user.id),
    )
    return ConversationHandler.END


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    global last_conflict_warning_at

    if is_polling_conflict(context.error):
        current_time = time.monotonic()
        if current_time - last_conflict_warning_at > 60:
            logger.warning("Xatolik: %s", format_runtime_error(context.error))
            last_conflict_warning_at = current_time
        return

    logger.error("Xatolik: %s", format_runtime_error(context.error))

    if isinstance(update, Update) and update.effective_message:
        await update.effective_message.reply_text(
            "Kechirasiz, texnik xatolik yuz berdi. Iltimos, qayta urinib ko'ring."
        )


async def on_bot_ready(application: Application) -> None:
    logger.info("Bot ishga tushdi.")


def is_polling_conflict(error: Exception | None) -> bool:
    if error is None:
        return False

    text = str(error).lower()
    return "terminated by other getupdates" in text or (
        "conflict" in text and "getupdates" in text
    )


def get_webhook_base_url() -> str | None:
    if WEBHOOK_URL:
        return WEBHOOK_URL.rstrip("/")
    if RAILWAY_PUBLIC_DOMAIN:
        return f"https://{RAILWAY_PUBLIC_DOMAIN}".rstrip("/")
    return None


def run_bot(application: Application) -> None:
    webhook_base_url = get_webhook_base_url()
    if webhook_base_url:
        webhook_url = f"{webhook_base_url}/{WEBHOOK_PATH}"
        logger.info("Bot webhook rejimida ishga tushmoqda.")
        application.run_webhook(
            listen="0.0.0.0",
            port=PORT,
            url_path=WEBHOOK_PATH,
            webhook_url=webhook_url,
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True,
        )
        return

    logger.info("Bot polling rejimida ishga tushmoqda.")
    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
    )


def main() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN .env faylida ko'rsatilmagan.")
    if not ADMIN_CHAT_ID:
        raise RuntimeError("ADMIN_CHAT_ID .env faylida ko'rsatilmagan.")
    if not ADMIN_IDS:
        logger.warning("ADMIN_IDS bo'sh. Admin panel hech kimga chiqmasligi mumkin.")

    init_db()
    application = Application.builder().token(BOT_TOKEN).post_init(on_bot_ready).build()

    application_conversation = ConversationHandler(
        entry_points=[
            CommandHandler("start", start),
            MessageHandler(filters.Regex(r"^(📝\s*)?Ariza yuborish$"), begin_application),
            CallbackQueryHandler(subscription_check_callback, pattern=r"^sub:check$"),
        ],
        states={
            FULL_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, full_name)],
            BIRTH_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, birth_date)],
            ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, address)],
            BRANCH: [MessageHandler(filters.TEXT & ~filters.COMMAND, branch)],
            EDUCATION: [
                CallbackQueryHandler(education_choice, pattern=r"^education:(higher|secondary)$"),
                MessageHandler(filters.ALL, invalid_education),
            ],
            JOB_DIRECTION: [
                CallbackQueryHandler(job_direction_choice, pattern=r"^job:(admin|teacher|assistant)$"),
                MessageHandler(filters.ALL, invalid_job_direction),
            ],
            SPECIALTY: [
                CallbackQueryHandler(specialty_choice, pattern=r"^specialty:(math|physics|russian|english)$"),
                MessageHandler(filters.ALL, invalid_specialty),
            ],
            EXPERIENCE_CHOICE: [
                CallbackQueryHandler(experience_choice_callback, pattern=r"^exp:(yes|no)$"),
                MessageHandler(filters.ALL, invalid_experience_choice),
            ],
            EXPERIENCE_YEARS: [MessageHandler(filters.TEXT & ~filters.COMMAND, experience_years)],
            CERTIFICATE_CHOICE: [
                CallbackQueryHandler(certificate_choice_callback, pattern=r"^cert:(yes|no)$"),
                MessageHandler(filters.ALL, invalid_certificate_choice),
            ],
            CERTIFICATE_PHOTO: [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE, certificate_photo),
                MessageHandler(filters.ALL, invalid_certificate_photo),
            ],
            PREVIOUS_JOB: [MessageHandler(filters.TEXT & ~filters.COMMAND, previous_job)],
            CONVICTED: [
                CallbackQueryHandler(convicted_choice, pattern=r"^convicted:(yes|no)$"),
                MessageHandler(filters.ALL, invalid_convicted),
            ],
            FAMILY_STATUS: [MessageHandler(filters.TEXT & ~filters.COMMAND, family_status)],
            PREVIOUS_SALARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, previous_salary)],
            EXPECTED_SALARY: [MessageHandler(filters.TEXT & ~filters.COMMAND, expected_salary)],
            WORD_LEVEL: [
                CallbackQueryHandler(word_level_choice, pattern=r"^word:(unknown|basic|medium|good)$"),
                MessageHandler(filters.ALL, invalid_word_level),
            ],
            EXCEL_LEVEL: [
                CallbackQueryHandler(excel_level_choice, pattern=r"^excel:(unknown|basic|medium|good)$"),
                MessageHandler(filters.ALL, invalid_excel_level),
            ],
            LANGUAGES: [MessageHandler(filters.TEXT & ~filters.COMMAND, languages)],
            FARIKS_DURATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, fariks_duration)],
            MOTIVATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, motivation)],
            PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, phone)],
            REVIEW_APPLICATION: [
                CallbackQueryHandler(review_callback, pattern=r"^review:(confirm|edit|cancel)$"),
                CallbackQueryHandler(edit_field_callback, pattern=r"^edit:[a-z_]+$"),
                CallbackQueryHandler(edit_value_callback, pattern=r"^editval:(education|convicted|word_level|excel_level|job_direction|specialty|certificate_status):[a-z]+$"),
                MessageHandler(filters.ALL, invalid_review_input),
            ],
            EDIT_FIELD: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_field_text)],
            RECENT_PHOTO: [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE, recent_photo),
                MessageHandler(filters.ALL, invalid_recent_photo),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", start),
        ],
    )

    application.add_handler(application_conversation)
    application.add_handler(
        CallbackQueryHandler(handle_application_decision, pattern=r"^app:(approve|reject):\d+$")
    )
    application.add_handler(CallbackQueryHandler(admin_add_start, pattern=r"^admin:add$"))
    application.add_handler(CallbackQueryHandler(admin_remove_start, pattern=r"^admin:remove$"))
    application.add_handler(CallbackQueryHandler(admin_excel_callback, pattern=r"^admin:excel$"))
    application.add_handler(CallbackQueryHandler(admin_status_list_callback, pattern=r"^admin:list:(pending|approved|rejected)$"))
    application.add_handler(CallbackQueryHandler(admin_view_application_callback, pattern=r"^admin:view:\d+$"))
    application.add_handler(CallbackQueryHandler(admin_search_start, pattern=r"^admin:search$"))
    application.add_handler(CallbackQueryHandler(admin_filter_start, pattern=r"^admin:filter_branch$"))
    application.add_handler(CallbackQueryHandler(admin_panel_callback, pattern=r"^admin:refresh$"))
    application.add_handler(CommandHandler("myid", my_id))
    application.add_handler(CommandHandler("cancel", admin_cancel))
    application.add_handler(MessageHandler(filters.Regex(r"^(🛠\s*)?Admin panel$"), admin_panel))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, admin_pending_text))
    application.add_error_handler(error_handler)
    try:
        run_bot(application)
    except Exception as error:
        logger.error("Botni ishga tushirishda xatolik: %s", format_runtime_error(error))
        sys.exit(1)


def format_runtime_error(error: Exception | None) -> str:
    if error is None:
        return "Noma'lum xatolik."

    text = str(error)
    lower_text = text.lower()

    if "unauthorized" in lower_text:
        return "Bot token noto'g'ri yoki bot o'chirilgan."
    if "not found" in lower_text and "chat" in lower_text:
        return "Admin yoki kanal ID topilmadi. ADMIN_CHAT_ID/PUBLISH_CHAT_ID ni tekshiring."
    if "forbidden" in lower_text:
        return "Botda ruxsat yo'q. Botni admin guruh/kanalga admin qiling yoki foydalanuvchi botni bloklagan."
    if "terminated by other getupdates" in lower_text or ("conflict" in lower_text and "getupdates" in lower_text):
        return (
            "Bot boshqa joyda ham ishlayapti. Bir xil token bilan faqat bitta bot ishlashi kerak: "
            "local `python main.py`ni to'xtating yoki Railway replicasini 1 ta qiling."
        )
    if "no module named" in lower_text:
        return "Kerakli kutubxona o'rnatilmagan. `pip install -r requirements.txt` qiling."

    return text


if __name__ == "__main__":
    main()
